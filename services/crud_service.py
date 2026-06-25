import base64
from io import BytesIO

from odoo.exceptions import AccessError, UserError, ValidationError

from ..utils.normalizer import ApiNormalizer
from ..utils.pdf_exporter import build_pdf
from ..utils.serializer import ApiSerializer


def get_model(env, config):
    return env[config["model"]].sudo()


def get_all(env, config, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    records = get_model(env, config).search([])
    return ApiSerializer(fields).records(records)


def get_by_page(env, config, page=1, size=20, column_list=None, keyword=None):
    page = max(int(page or 1), 1)
    size = max(int(size or 20), 1)
    normalizer = ApiNormalizer(config)
    fields = normalizer.columns(column_list)
    domain = normalizer.search_domain(keyword)
    model = get_model(env, config)
    total_record = model.search_count(domain)
    records = model.search(domain, offset=(page - 1) * size, limit=size)
    total_page = (total_record + size - 1) // size

    return {
        "items": ApiSerializer(fields).records(records),
        "current_page": page,
        "page_size": size,
        "total_record": total_record,
        "total_page": total_page,
    }


def get_by_id(env, config, record_id, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    record = get_model(env, config).browse(int(record_id)).exists()
    if not record:
        return None
    return ApiSerializer(fields).record(record)


def create_record(env, config, values):
    values = ApiNormalizer(config).payload(values)
    record = get_model(env, config).create(values)
    return ApiSerializer(config["fields"]).record(record)


def update_record(env, config, record_id, values):
    record = get_model(env, config).browse(int(record_id)).exists()
    if not record:
        return None
    values = ApiNormalizer(config).payload(values)
    record.write(values)
    return ApiSerializer(config["fields"]).record(record)


def copy_record(env, config, record_id, default=None):
    record = get_model(env, config).browse(int(record_id)).exists()
    if not record:
        return None
    default = default or {}
    for unique_field in ("code", "email", "username"):
        if unique_field in config["fields"] and unique_field not in default:
            value = record[unique_field]
            if unique_field == "email" and value and "@" in value:
                local_part, domain = value.split("@", 1)
                default[unique_field] = f"{local_part}.copy@{domain}"
            elif value:
                default[unique_field] = f"{value}-copy"
    new_record = record.copy(default)
    return ApiSerializer(config["fields"]).record(new_record)


def delete_record(env, config, record_id=None, ids=None):
    ids = ids or ApiNormalizer(config).ids(record_id=record_id)
    record = get_model(env, config).browse(ids).exists()
    if not record:
        return False
    deleted_count = len(record)
    record.unlink()
    return deleted_count


def template_import(config):
    return build_workbook(config["fields"], [])


def export_records(env, config, ids=None, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    model = get_model(env, config)
    domain = [("id", "in", ids)] if ids else []
    records = model.search(domain)
    rows = ApiSerializer(fields).rows(records)
    return build_workbook(fields, rows)


def export_pdf(env, config, resource, ids=None, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    model = get_model(env, config)
    domain = [("id", "in", ids)] if ids else []
    records = model.search(domain)
    rows = ApiSerializer(fields).rows(records)
    return build_pdf(resource, fields, rows)


def import_records(env, config, file_content):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UserError("Can cai thu vien openpyxl de import Excel.") from exc

    workbook = load_workbook(BytesIO(base64.b64decode(file_content)))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1] if cell.value]
    created = 0
    model = get_model(env, config)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and value is not None
        }
        if values:
            values = ApiNormalizer(config).payload(values)
            model.create(values)
            created += 1

    return {"created": created}


def build_workbook(headers, rows):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise UserError("Can cai thu vien openpyxl de export Excel.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def readable_error(exc):
    if isinstance(exc, (AccessError, UserError, ValidationError)):
        return str(exc)
    return "Co loi xay ra khi xu ly yeu cau."
