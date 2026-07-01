import base64
import re
from io import BytesIO

from odoo.exceptions import AccessError, UserError, ValidationError

from ..utils.normalizer import ApiNormalizer
from ..utils.pdf_exporter import build_pdf
from ..utils.serializer import ApiSerializer
from . import cache_service
from . import external_db_service
from .message_service import publish_event


TEXT_COPY_SUFFIX_RE = re.compile(r"(?i)(?:-copy(?:-\d+)?|-cp\d+)+$")
EMAIL_COPY_SUFFIX_RE = re.compile(r"(?i)(?:\.copy(?:-\d+)?|\.cp\d+)+$")
LEGACY_TEXT_COPY_RE = re.compile(r"(?i)-copy")
LEGACY_EMAIL_COPY_RE = re.compile(r"(?i)\.copy")


def get_model(env, config):
    return env[config["model"]].sudo()


def get_all(env, config, column_list=None):
    normalize_legacy_copy_values(env, config)
    fields = ApiNormalizer(config).columns(column_list)
    records = get_model(env, config).search([])
    return ApiSerializer(fields).records(records)


def get_by_page(env, config, page=1, size=20, column_list=None, keyword=None):
    normalize_legacy_copy_values(env, config)
    page = max(int(page or 1), 1)
    size = max(int(size or 20), 1)
    normalizer = ApiNormalizer(config)
    fields = normalizer.columns(column_list)
    domain = normalizer.search_domain(keyword)
    model = get_model(env, config)
    total_record = model.search_count(domain)
    records = model.search(domain, offset=(page - 1) * size, limit=size)
    total_page = (total_record + size - 1) // size

    return {
        "items": ApiSerializer(fields).records(records),
        "current_page": page,
        "page_size": size,
        "total_record": total_record,
        "total_page": total_page,
    }


def get_by_id(env, config, record_id, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    record = get_model(env, config).browse(int(record_id)).exists()
    if not record:
        return None
    return ApiSerializer(fields).record(record)


def create_record(env, config, values):
    values = ApiNormalizer(config).payload(values)
    record = get_model(env, config).create(values)
    external_db_service.sync_odoo_record(config["model"], record)
    cache_service.clear("external:")
    data = ApiSerializer(config["fields"]).record(record)
    publish_event(f"{config['model']}.created", data)
    return data


def update_record(env, config, record_id, values):
    record = get_model(env, config).browse(int(record_id)).exists()
    if not record:
        return None
    old_code = record.code if "code" in config["fields"] else None
    values = ApiNormalizer(config).payload(values)
    record.write(values)
    external_db_service.sync_odoo_record(config["model"], record, old_code=old_code)
    cache_service.clear("external:")
    data = ApiSerializer(config["fields"]).record(record)
    publish_event(f"{config['model']}.updated", data)
    return data


def copy_record(env, config, record_id, default=None):
    record = get_model(env, config).browse(int(record_id)).exists()
    if not record:
        return None
    default = default or {}
    for unique_field in ("code", "email", "username"):
        if unique_field in config["fields"] and unique_field not in default:
            value = record[unique_field]
            if unique_field == "email" and value and "@" in value:
                default[unique_field] = unique_email_copy_value(env, config, unique_field, value)
            elif value:
                default[unique_field] = unique_text_copy_value(env, config, unique_field, value)
    new_record = record.copy(default)
    external_db_service.sync_odoo_record(config["model"], new_record)
    cache_service.clear("external:")
    data = ApiSerializer(config["fields"]).record(new_record)
    publish_event(f"{config['model']}.copied", data)
    return data


def field_size(env, config, field_name):
    field = get_model(env, config)._fields.get(field_name)
    return getattr(field, "size", None)


def trim_to_size(value, max_size):
    if max_size and len(value) > max_size:
        return value[:max_size]
    return value


def copy_base_value(value, suffix_pattern):
    return suffix_pattern.sub("", str(value or "").strip())


def unique_text_copy_value(env, config, field_name, value):
    max_size = field_size(env, config, field_name)

    def make_candidate(index):
        suffix = f"-CP{index:02d}"
        base = copy_base_value(value, TEXT_COPY_SUFFIX_RE)
        if max_size:
            base = trim_to_size(base, max(max_size - len(suffix), 1))
        return f"{base}{suffix}"

    return unique_copy_value(env, config, field_name, make_candidate)


def unique_email_copy_value(env, config, field_name, value):
    max_size = field_size(env, config, field_name)
    local_part, domain = str(value).split("@", 1)

    def make_candidate(index):
        suffix = f".cp{index:02d}"
        base = copy_base_value(local_part, EMAIL_COPY_SUFFIX_RE)
        if max_size:
            max_local_size = max(max_size - len(domain) - 1 - len(suffix), 1)
            base = trim_to_size(base, max_local_size)
        return f"{base}{suffix}@{domain}"

    return unique_copy_value(env, config, field_name, make_candidate)


def unique_copy_value(env, config, field_name, make_candidate):
    model = get_model(env, config)
    index = 1
    candidate = make_candidate(index)
    while model.search_count([(field_name, "=", candidate)]):
        index += 1
        candidate = make_candidate(index)
    return candidate


def normalize_legacy_copy_values(env, config):
    unique_fields = [field for field in ("code", "email", "username") if field in config["fields"]]
    if not unique_fields:
        return

    model = get_model(env, config)
    records = model.search([])
    normalized = False
    for record in records:
        values = {}
        for field_name in unique_fields:
            value = record[field_name]
            if not value:
                continue
            if field_name == "email" and LEGACY_EMAIL_COPY_RE.search(value):
                values[field_name] = unique_email_copy_value(env, config, field_name, value)
            elif field_name != "email" and LEGACY_TEXT_COPY_RE.search(value):
                values[field_name] = unique_text_copy_value(env, config, field_name, value)

        if values:
            old_code = record.code if "code" in config["fields"] else None
            record.write(values)
            external_db_service.sync_odoo_record(config["model"], record, old_code=old_code)
            publish_event(f"{config['model']}.legacy_copy_normalized", ApiSerializer(config["fields"]).record(record))
            normalized = True

    if normalized:
        cache_service.clear("external:")


def copy_records(env, config, ids):
    copied = []
    for record_id in ids:
        data = copy_record(env, config, record_id)
        if data:
            copied.append(data)
    return copied


def delete_record(env, config, record_id=None, ids=None):
    ids = ids or ApiNormalizer(config).ids(record_id=record_id)
    record = get_model(env, config).browse(ids).exists()
    if not record:
        return False
    deleted_count = len(record)
    deleted_ids = record.ids
    external_db_service.delete_odoo_records(config["model"], record)
    cache_service.clear("external:")
    record.unlink()
    publish_event(f"{config['model']}.deleted", {"ids": deleted_ids, "deleted": deleted_count})
    return deleted_count


def template_import(config):
    return build_workbook(config["fields"], [])


def export_records(env, config, ids=None, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    model = get_model(env, config)
    domain = [("id", "in", ids)] if ids else []
    records = model.search(domain)
    rows = ApiSerializer(fields).rows(records)
    return build_workbook(fields, rows)


def export_pdf(env, config, resource, ids=None, column_list=None):
    fields = ApiNormalizer(config).columns(column_list)
    model = get_model(env, config)
    domain = [("id", "in", ids)] if ids else []
    records = model.search(domain)
    rows = ApiSerializer(fields).rows(records)
    return build_pdf(resource, fields, rows)


def import_records(env, config, file_content):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UserError("Can cai thu vien openpyxl de import Excel.") from exc

    workbook = load_workbook(BytesIO(base64.b64decode(file_content)))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1] if cell.value]
    created = 0
    model = get_model(env, config)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and value is not None
        }
        if values:
            values = ApiNormalizer(config).payload(values)
            record = model.create(values)
            external_db_service.sync_odoo_record(config["model"], record)
            cache_service.clear("external:")
            publish_event(f"{config['model']}.imported", ApiSerializer(config["fields"]).record(record))
            created += 1

    return {"created": created}


def build_workbook(headers, rows):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise UserError("Can cai thu vien openpyxl de export Excel.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def readable_error(exc):
    if isinstance(exc, (AccessError, UserError, ValidationError)):
        return str(exc)
    return "Co loi xay ra khi xu ly yeu cau."
